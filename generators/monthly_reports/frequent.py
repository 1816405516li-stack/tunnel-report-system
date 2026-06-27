"""Generate per-tunnel 机电经常性检查记录表."""

from __future__ import annotations

from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl.cell.cell import Cell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from generators.monthly_reports.common import (
    CATEGORY_LABELS,
    TEMPLATES,
    apply_generated_font,
    filter_by_tunnel,
    iter_tunnels,
    replace_cell_text,
    report_date_text,
    save_and_verify,
    set_generated_value,
    verify_workbook,
    short_tunnel_name,
    tunnel_dir,
    tunnel_display_name,
)


FREQUENT_WEEKLY_ROWS = (34, 176, 189, 199)
FREQUENT_WEEKLY_COLUMNS = (5, 6, 7, 8, 9)


def generate_frequent_reports(
    run_dir: Path,
    month: str,
    data: pd.DataFrame,
    manifest: pd.DataFrame,
    artifacts: dict[str, str],
    advance: Callable[[str], None],
) -> dict[str, int | str | bool]:
    """Create one final frequent inspection workbook per tunnel."""
    created: list[Path] = []
    row_count = 0

    for tunnel_name, tunnel_code in iter_tunnels(manifest, data):
        workbook = load_workbook(TEMPLATES["frequent"])
        sheet = workbook.active
        sheet.title = f"表5.机电经常性检查记录表({short_tunnel_name(tunnel_name)}隧道)"
        display_name = tunnel_display_name(tunnel_name)
        sheet["A2"].value = f"隧道名称：     {display_name}     （上行洞/下行洞）                路线名称：   修大高速  \xa0    "
        sheet["A3"].value = f"隧道编码：     {tunnel_code}                              路线编码：     S81     "
        note_spaces = "                                               " if tunnel_name == "小湖坳隧道" else "      "
        sheet["A4"].value = (
            f"养护机构：     永新东养护站           检查日期： {report_date_text(month, spaced=True)}"
            f"{note_spaces}注：正常为√，异常为×，异常且严重为△ ；周检以周一所在月"
        )
        for coordinate in ("A2", "A3", "A4"):
            apply_generated_font(sheet[coordinate], 10)

        tunnel_data = filter_by_tunnel(data, tunnel_name)
        row_count += len(tunnel_data)
        weekly_dates = weekly_dates_from_text(tunnel_data["周检日期"].iloc[0]) if not tunnel_data.empty else []
        fill_weekly_cells(sheet, weekly_dates)
        fill_frequent_abnormal(sheet, tunnel_data, tunnel_name)
        if tunnel_name == "高且隧道":
            normalize_frequent_merges(sheet, tunnel_name)
            delete_row_with_merges(sheet, 170)
            if "E170:I170" in {str(item) for item in sheet.merged_cells.ranges}:
                sheet.unmerge_cells("E170:I170")
            if "E170:G170" not in {str(item) for item in sheet.merged_cells.ranges}:
                sheet.merge_cells("E170:G170")
            if "H170:I170" not in {str(item) for item in sheet.merged_cells.ranges}:
                sheet.merge_cells("H170:I170")
            sheet._cells[(170, 8)] = Cell(sheet, row=170, column=8, value="√")
            apply_generated_font(sheet.cell(170, 8), 10)

        path = tunnel_dir(run_dir, tunnel_name, manifest) / "机电经常性检查记录表.xlsx"
        if tunnel_name == "高且隧道":
            path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(path)
            verify_workbook(path)
        else:
            save_and_verify(workbook, path, TEMPLATES["frequent"], "frequent")
            saved_workbook = load_workbook(path)
            normalize_frequent_merges(saved_workbook.active, tunnel_name)
            saved_workbook.save(path)
            verify_workbook(path)
        created.append(path)
        advance(f"{CATEGORY_LABELS['frequent']}：{tunnel_name} 生成完成")

    artifacts["frequent"] = str(run_dir)
    return {"label": CATEGORY_LABELS["frequent"], "file_count": len(created), "row_count": int(row_count), "ready": True}


def fill_frequent_abnormal(sheet: Worksheet, data: pd.DataFrame, tunnel_name: str) -> None:
    """Fill abnormal description and measure columns from 25th-day records."""
    abnormal = data[data["是否无异常"].astype(str) != "True"] if not data.empty and "是否无异常" in data.columns else data
    set_generated_value(sheet["L7"], "工单/照片/附件" if tunnel_name == "石桥二号隧道" else "", 10)
    if abnormal.empty:
        set_generated_value(sheet["J7"], "", 10)
        set_generated_value(sheet["K7"], "", 10)
        return
    set_generated_value(sheet["J7"], "\n".join(str(value) for value in abnormal["异常描述"].tolist() if str(value)), 10)
    set_generated_value(sheet["K7"], "\n".join(str(value) for value in abnormal["养护措施"].tolist() if str(value)), 10)


def fill_weekly_cells(sheet: Worksheet, weekly_dates: list[str]) -> None:
    """Fill all weekly inspection mini-grids used by the frequent template."""
    for row in (50, 51, 64, 65, 79, 80, 94, 95, 108, 109):
        for column in FREQUENT_WEEKLY_COLUMNS:
            set_generated_value(sheet.cell(row, column), "/", 10)
    set_generated_value(sheet["K133"], "", 10)
    for row in FREQUENT_WEEKLY_ROWS:
        for index, column in enumerate(FREQUENT_WEEKLY_COLUMNS):
            value = weekly_dates[index] if index < len(weekly_dates) else "/"
            set_generated_value(sheet.cell(row, column), value, 10)
            set_generated_value(sheet.cell(row + 1, column), "√" if index < len(weekly_dates) else "/", 10)


def normalize_frequent_merges(sheet: Worksheet, tunnel_name: str) -> None:
    """Apply common merge corrections found in the final frequent-inspection template."""
    existing = {str(item) for item in sheet.merged_cells.ranges}
    for range_text in (
        "D169:D170",
        "J169:J171",
        "K169:K171",
        "L169:L171",
        "J170:J171",
        "K170:K171",
        "L170:L171",
    ):
        if range_text not in {str(item) for item in sheet.merged_cells.ranges}:
            continue
        sheet.unmerge_cells(range_text)

    merge_profiles = {
        "石桥二号隧道": ("D169:D170",),
        "永井隧道": ("D169:D170",),
        "小湖坳隧道": ("J169:J171", "K169:K171", "L169:L171"),
        "梨树下隧道": ("D169:D170", "J169:J171", "K169:K171", "L169:L171"),
        "高且隧道": ("J169:J171", "K169:K171", "L169:L171"),
        "肖家山隧道": ("D169:D170", "J170:J171", "K170:K171", "L170:L171"),
        "黄坳隧道": ("D169:D170", "J169:J171", "K169:K171", "L169:L171"),
        "横垄岗隧道": ("D169:D170", "J169:J171", "K169:K171", "L169:L171"),
    }
    for range_text in merge_profiles.get(tunnel_name, ()):
        if range_text not in {str(item) for item in sheet.merged_cells.ranges}:
            sheet.merge_cells(range_text)


def delete_row_with_merges(sheet: Worksheet, row_index: int) -> None:
    """Delete one row and rebuild merged ranges so template blocks keep their shape."""
    merged_ranges = [str(item) for item in sheet.merged_cells.ranges]
    for range_text in merged_ranges:
        sheet.unmerge_cells(range_text)
    sheet.delete_rows(row_index, 1)
    for range_text in merged_ranges:
        new_range = _shift_range_after_row_delete(range_text, row_index)
        if new_range:
            sheet.merge_cells(new_range)


def _shift_range_after_row_delete(range_text: str, row_index: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    if max_row < row_index:
        return range_text
    if min_row > row_index:
        min_row -= 1
        max_row -= 1
    elif min_row <= row_index <= max_row:
        max_row -= 1
    if max_row < min_row:
        return ""
    if min_row == max_row and min_col == max_col:
        return ""
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return f"{start}:{end}"


def weekly_dates_from_text(value) -> list[str]:
    return [part.strip() for part in str(value or "").split("、") if part.strip()]
