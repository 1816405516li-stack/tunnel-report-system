"""Generate per-tunnel 机电日常巡查记录表."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from generators.monthly_reports.common import (
    CATEGORY_LABELS,
    TEMPLATES,
    days_in_month,
    delete_rows_after,
    filter_by_tunnel,
    generation_progress_message,
    iter_tunnels,
    numbered_lines,
    save_and_verify,
    set_daily_title,
    set_generated_value,
    short_tunnel_name,
    tunnel_dir,
)


DAILY_CATEGORY_ROWS = {
    "供配电设施": 4,
    "照明设施": 5,
    "通风设施": 6,
    "消防设施": 7,
    "消防与救援设施": 7,
    "监控与通信设施": 8,
}


def generate_daily_reports(
    run_dir: Path,
    month: str,
    data: pd.DataFrame,
    manifest: pd.DataFrame,
    artifacts: dict[str, str],
    advance: Callable[[str], None],
) -> dict[str, int | str | bool]:
    """Create one final daily inspection workbook per tunnel."""
    created: list[Path] = []
    row_count = 0
    days = days_in_month(month)

    for tunnel_name, _tunnel_code in iter_tunnels(manifest, data):
        workbook = load_workbook(TEMPLATES["daily"])
        template_sheet = workbook.worksheets[0]
        while len(workbook.worksheets) > 1:
            workbook.remove(workbook.worksheets[-1])

        tunnel_data = filter_by_tunnel(data, tunnel_name)
        row_count += len(tunnel_data)
        for day_index, current_day in enumerate(days):
            sheet = template_sheet if day_index == 0 else workbook.copy_worksheet(template_sheet)
            sheet.title = f"{current_day.month}月{current_day.day}日"
            set_daily_title(
                sheet,
                "A1",
                "吉安西管理中心 ",
                short_tunnel_name(tunnel_name),
                " 隧道机电设施日常检查记录表",
            )
            set_generated_value(sheet["B2"], datetime(current_day.year, current_day.month, current_day.day), 16)
            fill_daily_sheet(sheet, tunnel_data, current_day)
            if current_day.day == 8:
                set_generated_value(sheet["E18"], "v", 16)

        path = tunnel_dir(run_dir, tunnel_name, manifest) / "机电日常巡查记录表.xlsx"
        save_and_verify(workbook, path, TEMPLATES["daily"], "daily")
        created.append(path)
        advance(generation_progress_message("daily", tunnel_name))

    artifacts["daily"] = str(run_dir)
    return {"label": CATEGORY_LABELS["daily"], "file_count": len(created), "row_count": int(row_count), "ready": True}


def fill_daily_sheet(sheet: Worksheet, tunnel_data: pd.DataFrame, current_day: date) -> None:
    """Fill one copied daily inspection sheet with same-day faults."""
    day_text = current_day.isoformat()
    day_rows = tunnel_data[tunnel_data["检查日期"].astype(str) == day_text] if not tunnel_data.empty else pd.DataFrame()
    if day_rows.empty:
        return
    for category, group in day_rows.groupby("检查项目", sort=False):
        target_row = DAILY_CATEGORY_ROWS.get(str(category), 8)
        descriptions = group["养护单位检查情况描述"].astype(str).tolist()
        measures = group["采取措施"].astype(str).tolist()
        set_generated_value(sheet.cell(target_row, 5), "正常(   )\n\n异常( √ )", 16)
        set_generated_value(sheet.cell(target_row, 6), numbered_lines(descriptions), 16)
        set_generated_value(sheet.cell(target_row, 7), numbered_lines(measures), 16)
