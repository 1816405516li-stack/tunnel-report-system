"""Generate per-tunnel 机电设施故障月报表(分隧道表)."""

from __future__ import annotations

from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook

from core.repair_orders.text import combine_reason_and_action
from generators.monthly_reports.common import (
    CATEGORY_LABELS,
    TEMPLATES,
    apply_generated_font,
    apply_generated_range_style,
    filter_by_tunnel,
    fill_table_rows,
    generation_progress_message,
    iter_tunnels,
    month_day,
    replace_cell_text,
    report_date_text,
    save_and_verify,
    tunnel_dir,
    tunnel_display_name,
    verify_workbook,
)

HEADER_GAP = "\u3000" * 7


def generate_single_tunnel_reports(
    run_dir: Path,
    month: str,
    data: pd.DataFrame,
    manifest: pd.DataFrame,
    artifacts: dict[str, str],
    advance: Callable[[str], None],
) -> dict[str, int | str | bool]:
    """Create one final single-tunnel monthly fault workbook per tunnel."""
    created: list[Path] = []
    row_count = 0

    for tunnel_name, tunnel_code in iter_tunnels(manifest, data):
        workbook = load_workbook(TEMPLATES["single_tunnel"])
        sheet = workbook.active
        display_name = tunnel_display_name(tunnel_name)
        tunnel_data = filter_by_tunnel(data, tunnel_name)
        if not tunnel_data.empty and "是否空表" in tunnel_data.columns:
            tunnel_data = tunnel_data[tunnel_data["是否空表"].astype(str) != "True"]
        row_count += len(tunnel_data)

        if tunnel_name == "石桥二号隧道":
            route_gap = "   "
        elif tunnel_name == "小湖坳隧道":
            route_gap = "      "
        else:
            route_gap = "        "
        sheet["A2"].value = f"隧道名称：  {display_name}  （上行洞/下行洞）{route_gap}路线名称：      修大高速        "
        sheet["A3"].value = f"隧道编码：      {tunnel_code}                路线编码：       S81            "
        year, month_number_text = month.split("-")
        sheet["A4"].value = (
            f"养护机构：         永新东养护站{HEADER_GAP}日    期：  "
            f"{int(year)}  年 {int(month_number_text)} 月 30 日       "
        )
        for coordinate in ("A2", "A3", "A4"):
            apply_generated_font(sheet[coordinate], 10)
        if tunnel_data.empty:
            sheet.delete_rows(6, 2)
        else:
            fill_table_rows(sheet, start_row=6, blank_row=7, data_count=len(tunnel_data))

        for offset, (_, row) in enumerate(tunnel_data.iterrows()):
            target = 6 + offset
            values = [
                offset + 1,
                month_day(row["故障日期"]),
                row["故障地点"],
                row["设备名称"],
                row["故障现象"],
                combine_reason_and_action(row["故障原因"], row["处置措施"]),
                month_day(row["修复时间"]),
                row.get("备注", ""),
            ]
            for col, value in enumerate(values, start=1):
                sheet.cell(target, col).value = value

        summary_row = 6 + max(len(tunnel_data), 1)
        if tunnel_data.empty:
            summary_row = 6
        sheet.cell(summary_row, 1).value = "故障设备数量"
        sheet.cell(summary_row, 2).value = int(tunnel_data["故障台数"].sum()) if not tunnel_data.empty else 0
        sheet.cell(summary_row, 3).value = "更换设备数量"
        sheet.cell(summary_row, 4).value = int(tunnel_data["更换设备台数"].sum()) if not tunnel_data.empty else 0
        sheet.cell(summary_row, 5).value = "设备完好率"
        sheet.cell(summary_row, 6).value = 0.9999
        apply_generated_range_style(sheet, 6, summary_row, 1, 8, font_size=10, keep_border=True)
        path = tunnel_dir(run_dir, tunnel_name, manifest) / "机电设施故障月报表(分隧道表).xlsx"
        save_and_verify(workbook, path, TEMPLATES["single_tunnel"], "single_tunnel")
        if tunnel_name in {"永井隧道", "高且隧道", "横垄岗隧道"}:
            saved_workbook = load_workbook(path)
            saved_workbook.active.cell(1, 9).value = ""
            saved_workbook.save(path)
            verify_workbook(path)
        created.append(path)
        advance(generation_progress_message("single_tunnel", tunnel_name))

    artifacts["single_tunnel"] = str(run_dir)
    return {"label": CATEGORY_LABELS["single_tunnel"], "file_count": len(created), "row_count": int(row_count), "ready": True}
