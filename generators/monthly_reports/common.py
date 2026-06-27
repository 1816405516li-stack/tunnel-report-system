"""Shared helpers for template-based monthly workbook generation."""

from __future__ import annotations

from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from config.settings import RESOURCES_DIR
from generators.monthly_reports.common_paths import TEMPLATE_BY_PROFILE
from generators.monthly_reports.template_validator import assert_monthly_template_valid


DATASET_KEYS = {
    "total": "机电设施故障月报表_总表_数据",
    "single_tunnel": "机电设施故障月报表_分隧道表_数据",
    "daily": "机电日常巡查记录表_数据",
    "frequent": "机电经常性检查记录表_数据",
    "fault_record": "隧道机电设备故障记录单_数据",
    "tunnel_manifest": "隧道清单",
}

CATEGORY_LABELS = {
    "total": "总月报表",
    "single_tunnel": "单隧道月报表",
    "daily": "日常巡查记录表",
    "frequent": "经常性检查记录表",
    "fault_record": "设备故障记录单",
}

DEFAULT_CATEGORIES = ("total", "single_tunnel", "daily", "frequent", "fault_record")
GROUPED_CATEGORIES = {"single_tunnel", "daily", "frequent", "fault_record"}
TUNNEL_DISPLAY_NAMES = {
    "小湖坳隧道": "遂川小湖坳隧道",
}

TEMPLATE_DIR = RESOURCES_DIR / "templates" / "monthly"
TEMPLATES = {
    "total": TEMPLATE_BY_PROFILE["total"],
    "single_tunnel": TEMPLATE_BY_PROFILE["single_tunnel"],
    "daily": TEMPLATE_BY_PROFILE["daily"],
    "frequent": TEMPLATE_BY_PROFILE["frequent"],
    "fault_record": TEMPLATE_BY_PROFILE["fault_record"],
}

CHINESE_FONT = "宋体"
WESTERN_FONT = "Times New Roman"
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


class MonthlyGenerationError(RuntimeError):
    """Raised when report generation cannot continue with the current data."""


def ensure_templates_exist(categories: Iterable[str]) -> None:
    missing = [str(TEMPLATES[category]) for category in categories if category in TEMPLATES and not TEMPLATES[category].exists()]
    if missing:
        raise MonthlyGenerationError("缺少长期复用模板：" + "；".join(missing))


def read_csv(path: str | Path) -> pd.DataFrame:
    source = Path(path)
    if not source.exists():
        raise MonthlyGenerationError(f"找不到中间数据文件：{source}")
    try:
        return pd.read_csv(source, encoding="utf-8-sig").fillna("")
    except Exception as exc:
        raise MonthlyGenerationError(f"读取中间数据失败：{source.name}，{exc}") from exc


def iter_tunnels(manifest: pd.DataFrame, data: pd.DataFrame) -> list[tuple[str, str]]:
    if not manifest.empty and {"隧道名称", "隧道编码"}.issubset(manifest.columns):
        return [(str(row["隧道名称"]), str(row["隧道编码"])) for _, row in manifest.iterrows()]
    names = sorted(str(name) for name in data["隧道名称"].dropna().unique().tolist()) if "隧道名称" in data else []
    return [(name, "") for name in names]


def filter_by_tunnel(data: pd.DataFrame, tunnel_name: str) -> pd.DataFrame:
    if data.empty or "隧道名称" not in data.columns:
        return pd.DataFrame(columns=data.columns)
    return data[data["隧道名称"].astype(str) == tunnel_name].copy()


def tunnel_prefix(tunnel_name: str, manifest: pd.DataFrame) -> str:
    if manifest.empty or "隧道名称" not in manifest.columns:
        return "0"
    names = [str(name) for name in manifest["隧道名称"].tolist()]
    return str(names.index(tunnel_name) + 1) if tunnel_name in names else "0"


def tunnel_dir(root: Path, tunnel_name: str, manifest: pd.DataFrame) -> Path:
    return root / f"{tunnel_prefix(tunnel_name, manifest)}-{tunnel_display_name(tunnel_name)}"


def save_and_verify(workbook, path: Path, template_path: Path | None = None, profile: str | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    verify_workbook(path)
    if template_path and profile:
        try:
            assert_monthly_template_valid(template_path, path, profile)
        except ValueError as exc:
            raise MonthlyGenerationError(str(exc)) from exc


def verify_workbook(path: Path) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        workbook.close()
    except Exception as exc:
        raise MonthlyGenerationError(f"生成的 Excel 无法打开：{path.name}，{exc}") from exc


def iter_packable_files(root: Path, exclude: set[Path] | None = None) -> Iterable[Path]:
    exclude = {path.resolve() for path in (exclude or set())}
    for path in root.rglob("*"):
        if path.is_file() and path.name != "generation_result.json" and path.resolve() not in exclude:
            yield path


def fill_table_rows(sheet: Worksheet, start_row: int, blank_row: int, data_count: int) -> None:
    count = max(data_count, 1)
    if count > 1:
        sheet.insert_rows(blank_row, count - 1)
        for row_index in range(start_row + 1, start_row + count):
            copy_row_style(sheet, start_row, row_index)
    sheet.delete_rows(start_row + count, 1)


def unmerge_rows(sheet: Worksheet, start_row: int, end_row: int) -> None:
    """Remove template merges that intersect a data row interval."""
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_row <= end_row:
            sheet.unmerge_cells(str(merged_range))


def unmerge_columns(sheet: Worksheet, start_column: int, end_column: int) -> None:
    """Remove template merges that intersect a column interval."""
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_col >= start_column and merged_range.min_col <= end_column:
            sheet.unmerge_cells(str(merged_range))


def delete_rows_after(sheet: Worksheet, last_row: int) -> None:
    """Remove trailing template rows after the final visible row."""
    if sheet.max_row > last_row:
        sheet.delete_rows(last_row + 1, sheet.max_row - last_row)


def trim_sheet_bounds(sheet: Worksheet, max_row: int, max_column: int) -> None:
    """Drop cells outside the intended visible bounds."""
    for key in list(sheet._cells):
        row, column = key
        if row > max_row or column > max_column:
            del sheet._cells[key]


def copy_row_style(sheet: Worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def set_generated_value(cell: Cell, value, font_size: int | float = 10) -> None:
    """Write generated content with reusable Chinese/Western font rules."""
    cell.font = _font_for_value(value, font_size)
    cell.value = _rich_text_value(value, font_size)


def apply_generated_font(cell: Cell, font_size: int | float = 10) -> None:
    value = cell.value
    cell.font = _font_for_value(value, font_size)
    cell.value = _rich_text_value(value, font_size)


def apply_generated_range_style(
    sheet: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    font_size: int | float = 10,
    keep_border: bool = False,
) -> None:
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            apply_generated_font(cell, font_size)
            if keep_border:
                cell.border = _complete_border(cell.border)
            cell.alignment = copy(cell.alignment or Alignment())
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical=cell.alignment.vertical or "center",
                text_rotation=cell.alignment.text_rotation,
                wrap_text=True if cell.alignment.wrap_text is None else cell.alignment.wrap_text,
                shrink_to_fit=cell.alignment.shrink_to_fit,
                indent=cell.alignment.indent,
            )


def set_daily_title(sheet: Worksheet, coordinate: str, prefix: str, tunnel_name: str, suffix: str) -> None:
    sheet[coordinate].value = CellRichText(
        TextBlock(InlineFont(rFont=CHINESE_FONT, sz=16), prefix),
        TextBlock(InlineFont(rFont=CHINESE_FONT, sz=36), tunnel_name),
        TextBlock(InlineFont(rFont=CHINESE_FONT, sz=16), suffix),
    )
    sheet[coordinate].font = Font(name=CHINESE_FONT, size=16)


def _rich_text_value(value, font_size: int | float):
    if not isinstance(value, str) or value.startswith("=") or value == "":
        return value
    blocks = []
    current = value[0]
    current_is_chinese = _is_chinese(value[0])
    for char in value[1:]:
        is_chinese = _is_chinese(char)
        if is_chinese == current_is_chinese:
            current += char
            continue
        blocks.append(_text_block(current, current_is_chinese, font_size))
        current = char
        current_is_chinese = is_chinese
    blocks.append(_text_block(current, current_is_chinese, font_size))
    return CellRichText(*blocks)


def _text_block(text: str, is_chinese: bool, font_size: int | float) -> TextBlock:
    font_name = CHINESE_FONT if is_chinese else WESTERN_FONT
    return TextBlock(InlineFont(rFont=font_name, sz=font_size), text)


def _font_for_value(value, font_size: int | float) -> Font:
    if isinstance(value, str) and value and not value.startswith("=") and _contains_chinese(value):
        return Font(name=CHINESE_FONT, size=font_size)
    return Font(name=WESTERN_FONT, size=font_size)


def _contains_chinese(value: str) -> bool:
    return any(_is_chinese(char) for char in value)


def _is_chinese(char: str) -> bool:
    return "\u4e00" <= char <= "\u9fff"


def _has_border(border: Border) -> bool:
    return any(getattr(side, "style", None) for side in (border.left, border.right, border.top, border.bottom))


def _complete_border(border: Border) -> Border:
    return Border(
        left=copy(border.left if getattr(border.left, "style", None) else THIN_BORDER.left),
        right=copy(border.right if getattr(border.right, "style", None) else THIN_BORDER.right),
        top=copy(border.top if getattr(border.top, "style", None) else THIN_BORDER.top),
        bottom=copy(border.bottom if getattr(border.bottom, "style", None) else THIN_BORDER.bottom),
        diagonal=copy(border.diagonal),
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=copy(border.vertical),
        horizontal=copy(border.horizontal),
    )


def replace_cell_text(sheet: Worksheet, coordinate: str, marker: str, value: str, occurrence: int = 1) -> None:
    current = str(sheet[coordinate].value or "")
    if marker not in current:
        sheet[coordinate].value = value
        return
    pieces = current.split(marker)
    output = []
    replaced = False
    for index, piece in enumerate(pieces):
        output.append(piece)
        if index < len(pieces) - 1:
            if index + 1 == occurrence and not replaced:
                output.append(value)
                replaced = True
            else:
                output.append(marker)
    sheet[coordinate].value = "".join(output)


def days_in_month(month: str) -> list[date]:
    year, month_number = [int(part) for part in month.split("-")]
    current = date(year, month_number, 1)
    days: list[date] = []
    while current.month == month_number:
        days.append(current)
        current += timedelta(days=1)
    return days


def report_date_text(month: str, spaced: bool = False) -> str:
    last = days_in_month(month)[-1]
    return f"{last.year} 年 {last.month} 月 {last.day} 日" if spaced else f"{last.year}年{last.month}月{last.day}日"


def month_number(month: str) -> int:
    return int(month.split("-")[1])


def month_day(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return f"{parsed.month}月{parsed.day}日"


def chinese_datetime(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return f"{parsed.year}年{parsed.month}月{parsed.day}日{parsed.hour}时{parsed.minute}分"


def chinese_deadline_datetime(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return f"{parsed.year}年{parsed.month:02d}月{parsed.day:02d}日{parsed.hour:02d}时{parsed.minute:02d}分"


def numbered_lines(values: list[str]) -> str:
    values = [value for value in values if value]
    if len(values) <= 1:
        return values[0] if values else ""
    return "\n".join(f"{index}.{value}" for index, value in enumerate(values, start=1))


def int_value(value) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def truthy(value) -> bool:
    return str(value).strip().lower() in {"true", "1", "是", "yes"}


def short_tunnel_name(tunnel_name: str) -> str:
    return tunnel_name[:-2] if tunnel_name.endswith("隧道") else tunnel_name


def tunnel_display_name(tunnel_name: str) -> str:
    return TUNNEL_DISPLAY_NAMES.get(tunnel_name, tunnel_name)
