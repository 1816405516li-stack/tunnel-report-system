"""Validate generated workbooks against reusable monthly templates."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from generators.monthly_reports.common_paths import TEMPLATE_BY_PROFILE


@dataclass(frozen=True)
class TemplateValidationIssue:
    """One structural mismatch between a generated workbook and its template."""

    file: str
    sheet: str
    check: str
    expected: str
    actual: str

    def format(self) -> str:
        return f"{self.file}｜{self.sheet}｜{self.check}：应为 {self.expected}，实际 {self.actual}"


def validate_monthly_workbook(template_path: Path, generated_path: Path, profile: str) -> list[TemplateValidationIssue]:
    """Validate a generated workbook using a category-specific template profile."""
    template_wb = load_workbook(template_path, data_only=False)
    generated_wb = load_workbook(generated_path, data_only=False)
    try:
        if profile in {"total", "single_tunnel", "frequent"}:
            return _validate_single_sheet_workbook(template_wb.active, generated_wb.active, generated_path, profile)
        if profile == "daily":
            return _validate_repeated_sheet_workbook(template_wb.worksheets[0], generated_wb.worksheets, generated_path)
        if profile == "fault_record":
            return _validate_repeated_sheet_workbook(template_wb.worksheets[0], generated_wb.worksheets, generated_path)
        return [TemplateValidationIssue(str(generated_path), "-", "校验类型", "已知类型", profile)]
    finally:
        template_wb.close()
        generated_wb.close()


def assert_monthly_template_valid(template_path: Path, generated_path: Path, profile: str) -> None:
    """Raise ValueError if generated workbook breaks template structure."""
    issues = validate_monthly_workbook(template_path, generated_path, profile)
    if issues:
        preview = "\n".join(issue.format() for issue in issues[:12])
        extra = f"\n另有 {len(issues) - 12} 个问题" if len(issues) > 12 else ""
        raise ValueError(f"模板结构校验失败：\n{preview}{extra}")


def validate_monthly_output_tree(output_dir: Path) -> list[TemplateValidationIssue]:
    """Validate every generated workbook found under one monthly output directory."""
    issues: list[TemplateValidationIssue] = []
    for workbook_path in output_dir.rglob("*.xlsx"):
        profile = _profile_from_output_path(workbook_path)
        if not profile:
            continue
        issues.extend(validate_monthly_workbook(TEMPLATE_BY_PROFILE[profile], workbook_path, profile))
    return issues


def _profile_from_output_path(path: Path) -> str | None:
    text = str(path)
    if path.name == "机电设施故障月报表(总表).xlsx":
        return "total"
    if "单隧道月报表" in text:
        return "single_tunnel"
    if "日常巡查记录表" in text:
        return "daily"
    if "经常性检查记录表" in text:
        return "frequent"
    if "设备故障记录单" in text:
        return "fault_record"
    return None


def _validate_single_sheet_workbook(
    template_sheet: Worksheet,
    generated_sheet: Worksheet,
    generated_path: Path,
    profile: str,
) -> list[TemplateValidationIssue]:
    issues: list[TemplateValidationIssue] = []
    _compare_columns(template_sheet, generated_sheet, generated_path, issues)
    _compare_page_setup(template_sheet, generated_sheet, generated_path, issues)
    if profile == "frequent":
        _compare_rows(template_sheet, generated_sheet, generated_path, issues, range(1, template_sheet.max_row + 1))
        _compare_merged_ranges(template_sheet, generated_sheet, generated_path, issues)
    else:
        _compare_rows(template_sheet, generated_sheet, generated_path, issues, range(1, min(4, template_sheet.max_row) + 1))
        _compare_merged_ranges(template_sheet, generated_sheet, generated_path, issues, stable_max_row=4)
    return issues


def _validate_repeated_sheet_workbook(
    template_sheet: Worksheet,
    generated_sheets: Iterable[Worksheet],
    generated_path: Path,
) -> list[TemplateValidationIssue]:
    issues: list[TemplateValidationIssue] = []
    for generated_sheet in generated_sheets:
        _compare_columns(template_sheet, generated_sheet, generated_path, issues)
        _compare_rows(template_sheet, generated_sheet, generated_path, issues, range(1, template_sheet.max_row + 1))
        _compare_merged_ranges(template_sheet, generated_sheet, generated_path, issues)
        _compare_page_setup(template_sheet, generated_sheet, generated_path, issues)
    return issues


def _compare_columns(template_sheet: Worksheet, generated_sheet: Worksheet, path: Path, issues: list[TemplateValidationIssue]) -> None:
    if template_sheet.max_column != generated_sheet.max_column:
        issues.append(_issue(path, generated_sheet, "最大列数", template_sheet.max_column, generated_sheet.max_column))
    max_column = min(template_sheet.max_column, generated_sheet.max_column)
    for column in range(1, max_column + 1):
        template_width = template_sheet.column_dimensions[_column_letter(column)].width
        generated_width = generated_sheet.column_dimensions[_column_letter(column)].width
        if _norm_number(template_width) != _norm_number(generated_width):
            issues.append(_issue(path, generated_sheet, f"{_column_letter(column)}列宽", template_width, generated_width))


def _compare_rows(
    template_sheet: Worksheet,
    generated_sheet: Worksheet,
    path: Path,
    issues: list[TemplateValidationIssue],
    rows: Iterable[int],
) -> None:
    for row in rows:
        template_height = template_sheet.row_dimensions[row].height
        generated_height = generated_sheet.row_dimensions[row].height
        if _norm_number(template_height) != _norm_number(generated_height):
            issues.append(_issue(path, generated_sheet, f"{row}行高", template_height, generated_height))


def _compare_merged_ranges(
    template_sheet: Worksheet,
    generated_sheet: Worksheet,
    path: Path,
    issues: list[TemplateValidationIssue],
    stable_max_row: int | None = None,
) -> None:
    expected = _merged_range_set(template_sheet, stable_max_row)
    actual = _merged_range_set(generated_sheet, stable_max_row)
    if expected != actual:
        issues.append(_issue(path, generated_sheet, "合并单元格", sorted(expected), sorted(actual)))


def _compare_page_setup(template_sheet: Worksheet, generated_sheet: Worksheet, path: Path, issues: list[TemplateValidationIssue]) -> None:
    checks = {
        "打印方向": (template_sheet.page_setup.orientation, generated_sheet.page_setup.orientation),
        "纸张大小": (template_sheet.page_setup.paperSize, generated_sheet.page_setup.paperSize),
        "左边距": (template_sheet.page_margins.left, generated_sheet.page_margins.left),
        "右边距": (template_sheet.page_margins.right, generated_sheet.page_margins.right),
        "上边距": (template_sheet.page_margins.top, generated_sheet.page_margins.top),
        "下边距": (template_sheet.page_margins.bottom, generated_sheet.page_margins.bottom),
    }
    for label, (expected, actual) in checks.items():
        if _norm_number(expected) != _norm_number(actual):
            issues.append(_issue(path, generated_sheet, label, expected, actual))


def _merged_range_set(sheet: Worksheet, stable_max_row: int | None = None) -> set[str]:
    values = set()
    for cell_range in sheet.merged_cells.ranges:
        if stable_max_row is None or cell_range.max_row <= stable_max_row:
            values.add(str(cell_range))
    return values


def _issue(path: Path, sheet: Worksheet, check: str, expected, actual) -> TemplateValidationIssue:
    return TemplateValidationIssue(str(path), sheet.title, check, str(expected), str(actual))


def _column_letter(column: int) -> str:
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _norm_number(value) -> str:
    if value is None:
        return ""
    try:
        return f"{float(value):.6f}"
    except (TypeError, ValueError):
        return str(value)
