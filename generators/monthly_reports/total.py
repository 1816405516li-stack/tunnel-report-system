"""Generate 机电设施故障月报表(总表)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from generators.monthly_reports.common import (
    CATEGORY_LABELS,
    TEMPLATES,
    apply_generated_font,
    apply_generated_range_style,
    delete_rows_after,
    fill_table_rows,
    int_value,
    month_day,
    month_number,
    replace_cell_text,
    report_date_text,
    save_and_verify,
    unmerge_rows,
)

HEADER_GAP = "\u3000" * 7


def generate_total_report(
    run_dir: Path,
    month: str,
    data: pd.DataFrame,
    artifacts: dict[str, str],
) -> dict[str, int | str | bool]:
    """Create the final total monthly fault workbook from its template."""
    path = run_dir / "机电设施故障月报表(总表).xlsx"
    workbook = load_workbook(TEMPLATES["total"])
    sheet = workbook.active
    sheet.title = f"表3-1.机电设施故障月报表{month_number(month)}月"
    sheet["A3"].value = f"养护机构： 永新东养护站{HEADER_GAP}日    期： {report_date_text(month, spaced=True)}      "
    apply_generated_font(sheet["A3"], 10)
    fill_table_rows(sheet, start_row=5, blank_row=6, data_count=len(data))
    unmerge_rows(sheet, 5, 5 + max(len(data), 1) - 1)

    for offset, (_, row) in enumerate(data.iterrows()):
        target = 5 + offset
        values = [
            offset + 1,
            row["隧道名称"],
            f"{month_day(row['故障日期'])} ",
            row["故障地点"],
            row["设备分类"],
            row["设备名称"],
            row["故障现象"],
            row["故障原因"],
            row["处置措施"],
            f"{month_day(row['修复时间'])} ",
            row.get("备注", ""),
            int_value(row.get("更换设备台数")),
            int_value(row.get("故障台数")),
            int_value(row.get("故障天数")),
            f"=PRODUCT(M{target},N{target})",
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(target, col).value = value

    summary_row = 5 + max(len(data), 1)
    data_end_row = summary_row - 1
    sheet.cell(summary_row, 1).value = "故障设备数量"
    sheet.cell(summary_row, 2).value = f"=SUM(M5:M{data_end_row})"
    sheet.cell(summary_row, 4).value = "更换设备数量"
    sheet.cell(summary_row, 5).value = f"=SUM(L5:L{data_end_row})"
    sheet.cell(summary_row, 7).value = "设备完好率"
    sheet.cell(summary_row, 8).value = 0.9999
    sheet.cell(summary_row, 12).value = f"=SUM(L5:L{data_end_row})"
    sheet.cell(summary_row, 13).value = f"=SUM(M5:M{data_end_row})"
    sheet.cell(summary_row, 14).value = f"=SUM(N5:N{data_end_row})"
    sheet.cell(summary_row, 15).value = f"=SUM(O5:O{data_end_row})"
    apply_generated_range_style(sheet, 5, summary_row, 1, 15, font_size=10, keep_border=True)
    signature_row = summary_row + 1
    if f"A{signature_row}:K{signature_row}" not in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=11)
    delete_rows_after(sheet, summary_row + 1)
    save_and_verify(workbook, path, TEMPLATES["total"], "total")
    artifacts["total"] = str(path)
    return {"label": CATEGORY_LABELS["total"], "file_count": 1, "row_count": int(len(data)), "ready": True}
